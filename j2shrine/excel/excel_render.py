import openpyxl
from ..render import Render
from ..context import RenderContext
from .excel_custom_filter import excel_time

class ExcelRender(Render):

    # jinja2テンプレートの生成
    def __init__(self, *, context: RenderContext):
        super().__init__(context=context)
        self.cols = context.names.copy()

    def install_filters(self, *, environment):
        super().install_filters(environment=environment)
        environment.filters['excel_time'] = excel_time

    def build_reader(self, *, source: any):
        # 既にブックで渡された場合、そのまま返す
        if (isinstance(source, openpyxl.Workbook)):
            return source
        # ファイルの場合はロードする
        return openpyxl.load_workbook(source, data_only=True)

    def read_source(self, *, reader):

        cells = self.context.read_range
        sheets = self.context.sheets
        # endがnullの場合最後のシートまで読む
        if (sheets.end is None):
            sheets.end = len(reader.worksheets)-1

        results = []
        sheet_idx = sheets.start
        while sheet_idx <= sheets.end:
            sheet = reader.worksheets[sheet_idx]

            # コンテンツ読込み
            sheet_data = {
                'name': reader.sheetnames[sheet_idx],
                'rows': [],
                'abs': self.absolute_cells(sheet=sheet, cells=self.context.absolute)
            }
            for line_no, row in enumerate(sheet.iter_rows(min_col=cells.start.col, min_row=cells.start.row,
                                       max_col=cells.end.col, max_row=cells.end.row, )):
                sheet_data['rows'].append(self.read_row(line_no=line_no, columns=row))

            results.append(sheet_data)
            sheet_idx = 1 + sheet_idx
        return results

    # カラムのlistをdictに変換する。dictのキーはセル位置
    def read_row(self, *, line_no, columns):
        line = {}
        for index, column in enumerate(columns):
            letter = self.column_name(index)
            # カラムから値を取り出す
            line[letter] = self.read_column(name=letter, column=column)
        return line

    def absolute_cells(self, *, sheet, cells: dict):
        cell_values = {}
        for name, addr in cells.items():
            cell_values[name] = self.read_column(name=name, column=sheet[addr])
        return cell_values

    # openpyxlのCellオブジェクトからの値読み出し
    def read_column(self, *, name, column):
        # value属性が存在しない場合がある
        if (hasattr(column, 'value')):
            return column.value
        else:
            return None

    def get_column_letter(self, *, column):
        return openpyxl.utils.cell.get_column_letter(column.column)

    # カラム名取得
    def column_name(self, index):
        if len(self.cols) <= index:
            # カラム名が定義されていない場合
            # または定義済みのカラム名よりも実際のカラムが多い場合はカラム名を追加で生成する
            self.cols.append(self.context.prefix + str(index).zfill(2))
        return self.cols[index]

    # jinja2へ渡す読み取り結果
    def finish(self, *, result):
        final_result = {
            'sheets': result,
            'params': self.context.parameters
        }
        return final_result
